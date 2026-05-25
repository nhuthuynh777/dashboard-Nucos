"""
parse_data.py — Nucos Dashboard data parser
Reads a single xlsx file with multiple sheets:
  - Media Plan - Branding&Performan
  - Facebook Raw data
  - Google Raw data
  - Google KW Raw data
  - Messenger Sale Data
"""

import re
import datetime
import openpyxl
from collections import defaultdict
from io import BytesIO


# ── Helpers ──────────────────────────────────────────────────────────────────

def fv(v):
    """Parse numeric value. Handles European format (20.763 = 20,763)."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('₫', '').replace('%', '').replace('\xa0', '').strip()
    if not s or s == '-' or s == '--':
        return 0.0
    if re.match(r'^\d{1,3}(\.\d{3})+$', s):
        s = s.replace('.', '')
    else:
        s = s.replace(',', '')
    try:
        return float(s)
    except Exception:
        return 0.0


def sv(v):
    return str(v).strip() if v else ''


def short_name(ad_name):
    n = sv(ad_name)
    n = re.sub(r'^\d{2}/\d{2}\s*[-–|]\s*', '', n)
    return n[:60] + ('…' if len(n) > 60 else '')


def fv_count(v):
    """Parse Google Ads count (impressions/clicks).
    European thousands: 3.349 (float) = 3,349 actual. Cost/integers unchanged."""
    if v is None:
        return 0.0
    if isinstance(v, float):
        scaled = v * 1000
        if abs(scaled - round(scaled)) < 0.01:
            return float(round(scaled))
        return v
    return float(v) if isinstance(v, (int, float)) else fv(v)


def fmt_vnd(v):
    if v >= 1_000_000_000:
        return f"{v/1_000_000_000:.1f}B₫"
    if v >= 1_000_000:
        return f"{v/1_000_000:.1f}M₫"
    return f"{int(v):,}₫"


def fmt_num(v):
    if v >= 1_000_000:
        return f"{v/1_000_000:.1f}M"
    if v >= 1_000:
        return f"{v/1_000:.1f}K"
    return f"{int(v):,}"


# ── Classifiers ──────────────────────────────────────────────────────────────

def classify_campaign(camp):
    c = sv(camp).lower()
    if 'cpas' in c:
        return 'CPAS'
    if any(k in c for k in ('messag', 'messen', 'messenger', 'msg', 'mess', 'convers', 'tin nh')):
        return 'Messenger'
    if any(k in c for k in ('engage', 'engag', 'tương tác')):
        return 'Branding_Engage'
    if any(k in c for k in ('video', 'thruplay')):
        return 'Branding_Video'
    if 'reach' in c:
        return 'Branding_Reach'
    return 'Other'


def classify_targeting(adset):
    a = sv(adset)
    al = a.lower()
    if 'lal' in a.upper():
        return 'LAL'
    if 'retarget' in al:
        return 'Retargeting'
    if 'beauty brand' in al:
        return 'Beauty Brand'
    if 'sakura' in al:
        return 'Sakura Event'
    if 'cross' in al:
        return 'Cross Selling'
    if 'atc' in a or 'view content' in al:
        return 'ATC / View Content'
    if 'other cities' in al:
        return 'Prospecting – Other Cities'
    if '30+' in a:
        return 'Prospecting – HCM 30+'
    if 'hcm' in a.upper():
        return 'Prospecting – HCM'
    return 'Other / Broad'


# ── Sheet finder ──────────────────────────────────────────────────────────────

def find_sheet(wb, *keywords):
    for s in wb.sheetnames:
        sl = s.lower()
        if all(k.lower() in sl for k in keywords):
            return s
    return None


# ── Media Plan ────────────────────────────────────────────────────────────────

def parse_media_plan(ws):
    """
    Row 13=Reach, 14=Engage, 15=Video, 16=GDN, 17=Search, 19=Messenger.
    Col 7=Budget, 9=KPI count, 10=Unit cost, 12=Reach plan, 14=Impressions plan.
    """
    def r(row, col):
        return fv(ws.cell(row, col).value)

    def label(row_num):
        for col in range(1, 6):
            v = sv(ws.cell(row_num, col).value).lower()
            if v:
                return v
        return ''

    # Google rows 16-17: detect GDN vs Search by label
    row16_lbl = label(16); row17_lbl = label(17)
    gdn_row    = 16 if ('gdn' in row16_lbl or 'display' in row16_lbl or 'banner' in row16_lbl) else \
                 17 if ('gdn' in row17_lbl or 'display' in row17_lbl or 'banner' in row17_lbl) else 16
    search_row = 17 if gdn_row == 16 else 16

    google_gdn = {
        'budget': round(r(gdn_row, 7)),
        'clicks': round(r(gdn_row, 9)),
        'imp':    round(r(gdn_row, 14)),
    }
    google_search = {
        'budget':  round(r(search_row, 7)),
        'clicks':  round(r(search_row, 9)),
        'imp':     round(r(search_row, 14)),
        'ctr_plan': round(r(search_row, 9) / r(search_row, 14) * 100, 2) if r(search_row, 14) else 0,
    }

    # Total budget = Reach + Engage + Video + Messenger + GDN + Search
    reach_bud   = r(13, 7)
    engage_bud  = r(14, 7)
    video_bud   = r(15, 7)
    msg_bud     = r(19, 7)
    gdn_bud     = r(gdn_row, 7)
    search_bud  = r(search_row, 7)
    total_budget = reach_bud + engage_bud + video_bud + msg_bud + gdn_bud + search_bud

    return {
        'total_budget': round(total_budget),
        '_debug_budget': {
            'row13_reach_col7':      reach_bud,
            'row14_engage_col7':     engage_bud,
            'row15_video_col7':      video_bud,
            f'row{gdn_row}_gdn_col7':    gdn_bud,
            f'row{search_row}_search_col7': search_bud,
            'row19_messenger_col7':  msg_bud,
            'total':                 total_budget,
        },
        'reach': {
            'budget': round(r(13, 7)), 'cpm': round(r(13, 10)),
            'kpi':    round(r(13, 12)), 'imp': round(r(13, 14)),
        },
        'engage': {
            'budget':   round(r(14, 7)), 'cpe': round(r(14, 10)),
            'kpi':      round(r(14, 9)),  'imp': round(r(14, 14)),
            'ctr_plan': round(r(14, 21) * 100, 2),  # col U: CTR/ER/VTR → %
        },
        'video': {
            'budget':   round(r(15, 7)), 'cpv': round(r(15, 10)),
            'kpi':      round(r(15, 9)),  'imp': round(r(15, 14)),
            'vtr_plan': round(r(15, 21) * 100, 1),  # col U: VTR → %
        },
        'google_gdn':    google_gdn,
        'google_search': google_search,
        'messenger': {
            'budget':        round(r(19, 7)), 'cpmsg': round(r(19, 10)),
            'kpi':           r(19, 9),
            'purchase_plan': r(19, 22),   # col V19
            'gmv_plan':      round(r(19, 23)),  # col W19
        },
    }


# ── Facebook Raw ──────────────────────────────────────────────────────────────

def parse_fb_raw(ws):
    """
    Fixed column indices (0-based) per data logic doc:
    [2]=Ad name, [6]=Imp, [13]=CPAS purch, [16]=CPAS rev, [18]=Spend,
    [19]=Adset, [20]=Campaign, [21]=Clicks(all), [22]=Results, [24]=Reach,
    [25]=Msg started, [26]=Purchases, [27]=Purchase value, [28]=Msg replied
    """
    b_engage = defaultdict(lambda: dict(spend=0, imp=0, reach=0, clicks_all=0, results=0))
    b_video  = defaultdict(lambda: dict(spend=0, imp=0, reach=0, clicks_all=0, results=0))
    b_reach  = defaultdict(lambda: dict(spend=0, imp=0, reach=0, clicks_all=0, results=0))
    b_tgt    = defaultdict(lambda: dict(spend=0, imp=0, clicks_all=0))
    msg_camp = defaultdict(lambda: dict(spend=0, imp=0, clicks_all=0, msg=0, replied=0))
    msg_tgt  = defaultdict(lambda: dict(spend=0, msg=0, replied=0))
    msg_ads  = defaultdict(lambda: dict(spend=0, imp=0, clicks_all=0, msg=0, replied=0, purch=0, rev=0))
    cpas_tgt = defaultdict(lambda: dict(spend=0, purch=0, rev=0))
    cpas_ads = defaultdict(lambda: dict(spend=0, imp=0, clicks_all=0, purch=0, rev=0))
    unclassified = set()  # debug: campaign names that fell through to 'Other'

    def g(row, idx):
        return fv(row[idx]) if idx < len(row) else 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 23 or row[20] is None:
            continue
        camp  = sv(row[20])
        if not camp:
            continue
        typ   = classify_campaign(camp)
        ad    = sv(row[2])
        adset = sv(row[19])
        sp  = g(row, 18); imp = g(row, 6);  ca  = g(row, 21)
        res = g(row, 22); rch = g(row, 24)

        if typ == 'Branding_Engage':
            d = b_engage[ad]
            d['spend'] += sp; d['imp'] += imp; d['reach'] += rch
            d['clicks_all'] += ca; d['results'] += res
            t = b_tgt[classify_targeting(adset)]
            t['spend'] += sp; t['imp'] += imp; t['clicks_all'] += ca

        elif typ == 'Branding_Video':
            d = b_video[ad]
            d['spend'] += sp; d['imp'] += imp; d['reach'] += rch
            d['clicks_all'] += ca; d['results'] += res
            t = b_tgt[classify_targeting(adset)]
            t['spend'] += sp; t['imp'] += imp; t['clicks_all'] += ca

        elif typ == 'Branding_Reach':
            d = b_reach[ad]
            d['spend'] += sp; d['imp'] += imp; d['reach'] += rch
            d['clicks_all'] += ca; d['results'] += res
            t = b_tgt[classify_targeting(adset)]
            t['spend'] += sp; t['imp'] += imp; t['clicks_all'] += ca

        elif typ == 'Messenger':
            mc = msg_camp[camp]
            mc['spend'] += sp; mc['imp'] += imp; mc['clicks_all'] += ca
            mc['msg']     += g(row, 25)
            mc['replied'] += g(row, 28) if len(row) > 28 else 0
            mt = msg_tgt[classify_targeting(adset)]
            mt['spend'] += sp
            mt['msg']     += g(row, 25)
            mt['replied'] += g(row, 28) if len(row) > 28 else 0
            ma = msg_ads[ad]
            ma['spend'] += sp; ma['imp'] += imp; ma['clicks_all'] += ca
            ma['msg']     += g(row, 25)
            ma['replied'] += g(row, 28) if len(row) > 28 else 0
            ma['purch']   += g(row, 26) if len(row) > 26 else 0
            ma['rev']     += g(row, 27) if len(row) > 27 else 0

        elif typ == 'CPAS':
            ct = cpas_tgt[classify_targeting(adset)]
            ct['spend'] += sp
            ct['purch'] += g(row, 13)
            ct['rev']   += g(row, 16)
            ca_d = cpas_ads[ad]
            ca_d['spend'] += sp; ca_d['imp'] += imp; ca_d['clicks_all'] += ca
            ca_d['purch'] += g(row, 13)
            ca_d['rev']   += g(row, 16)

        else:
            unclassified.add(camp)

    def agg_brand(d):
        sp = sum(x['spend'] for x in d.values())
        imp = sum(x['imp'] for x in d.values())
        rch = sum(x['reach'] for x in d.values())
        ca  = sum(x['clicks_all'] for x in d.values())
        res = sum(x['results'] for x in d.values())
        ctr = round(ca / imp * 100, 2) if imp else 0.0
        return dict(spend=round(sp), imp=round(imp), reach=round(rch),
                    clicks_all=round(ca), results=round(res), ctr=ctr)

    engage_tot = agg_brand(b_engage)
    engage_tot['cpe'] = round(engage_tot['spend'] / engage_tot['results']) if engage_tot['results'] else 0
    video_tot = agg_brand(b_video)
    video_tot['cpv'] = round(video_tot['spend'] / video_tot['results']) if video_tot['results'] else 0
    reach_tot = agg_brand(b_reach)
    reach_tot['cpm'] = round(reach_tot['spend'] / reach_tot['imp'] * 1000) if reach_tot['imp'] else 0

    # Top engage ads
    engage_ads = []
    for rank, (name, d) in enumerate(sorted(b_engage.items(), key=lambda x: -x[1]['results'])[:5], 1):
        ctr = round(d['clicks_all'] / d['imp'] * 100, 2) if d['imp'] else 0
        cpe = round(d['spend'] / d['results']) if d['results'] else 0
        engage_ads.append(dict(rank=rank, name=name, short_name=short_name(name),
                               spend=round(d['spend']), imp=round(d['imp']),
                               results=round(d['results']), cpe=cpe, ctr=ctr))

    # Top video ads
    video_ads = []
    for rank, (name, d) in enumerate(sorted(b_video.items(), key=lambda x: -x[1]['results'])[:5], 1):
        ctr = round(d['clicks_all'] / d['imp'] * 100, 2) if d['imp'] else 0
        cpv = round(d['spend'] / d['results']) if d['results'] else 0
        video_ads.append(dict(rank=rank, name=name, short_name=short_name(name),
                              spend=round(d['spend']), imp=round(d['imp']),
                              results=round(d['results']), cpv=cpv, ctr=ctr))

    # Targeting table (Branding)
    tgt_rows = []
    for seg, d in sorted(b_tgt.items(), key=lambda x: -x[1]['spend']):
        ctr = round(d['clicks_all'] / d['imp'] * 100, 2) if d['imp'] else 0
        cpm = round(d['spend'] / d['imp'] * 1000) if d['imp'] else 0
        tgt_rows.append(dict(segment=seg, spend=round(d['spend']),
                             imp=round(d['imp']), ctr=ctr, cpm=cpm))

    # Messenger totals
    t_msg = sum(d['msg'] for d in msg_camp.values())
    t_rep = sum(d['replied'] for d in msg_camp.values())
    t_msp = sum(d['spend'] for d in msg_camp.values())
    msg_totals = dict(
        spend=round(t_msp), imp=round(sum(d['imp'] for d in msg_camp.values())),
        msg=round(t_msg), replied=round(t_rep),
        reply_rate=round(t_rep / t_msg * 100, 1) if t_msg else 0,
        cost_per_msg=round(t_msp / t_msg) if t_msg else 0
    )

    msg_camp_rows = []
    for name, d in sorted(msg_camp.items(), key=lambda x: -x[1]['msg']):
        rr   = round(d['replied'] / d['msg'] * 100, 1) if d['msg'] else 0
        cpm_v = round(d['spend'] / d['msg']) if d['msg'] else 0
        msg_camp_rows.append(dict(name=name, spend=round(d['spend']), imp=round(d['imp']),
                                  msg=round(d['msg']), replied=round(d['replied']),
                                  reply_rate=rr, cost_per_msg=cpm_v))

    msg_tgt_rows = []
    for seg, d in sorted(msg_tgt.items(), key=lambda x: -x[1]['msg']):
        cpm_v = round(d['spend'] / d['msg']) if d['msg'] else 0
        msg_tgt_rows.append(dict(segment=seg, spend=round(d['spend']),
                                 msg=round(d['msg']), replied=round(d['replied']),
                                 cost_per_msg=cpm_v))

    msg_ad_rows = []
    for rank, (name, d) in enumerate(sorted(msg_ads.items(), key=lambda x: -x[1]['msg'])[:8], 1):
        ctr   = round(d['clicks_all'] / d['imp'] * 100, 2) if d['imp'] else 0
        cpm_v = round(d['spend'] / d['msg']) if d['msg'] else 0
        msg_ad_rows.append(dict(rank=rank, name=name, short_name=short_name(name),
                                spend=round(d['spend']), msg=round(d['msg']),
                                cost_per_msg=cpm_v, purch=round(d['purch']),
                                rev=round(d['rev']), ctr=ctr))

    # CPAS totals
    cpas_sp = sum(d['spend'] for d in cpas_tgt.values())
    cpas_pu = sum(d['purch'] for d in cpas_tgt.values())
    cpas_rv = sum(d['rev']   for d in cpas_tgt.values())
    cpas_totals = dict(
        spend=round(cpas_sp), purch=round(cpas_pu), rev=round(cpas_rv),
        roas=round(cpas_rv / cpas_sp, 1) if cpas_sp else 0,
        cpa=round(cpas_sp / cpas_pu) if cpas_pu else 0
    )

    cpas_adset_rows = []
    for name, d in sorted(cpas_tgt.items(), key=lambda x: -x[1]['purch']):
        roas = round(d['rev'] / d['spend'], 1) if d['spend'] else 0
        cpa  = round(d['spend'] / d['purch']) if d['purch'] else 0
        cpas_adset_rows.append(dict(name=name, spend=round(d['spend']),
                                    purch=round(d['purch']), rev=round(d['rev']),
                                    roas=roas, cpa=cpa))

    cpas_ad_rows = []
    for rank, (name, d) in enumerate(sorted(cpas_ads.items(), key=lambda x: -x[1]['purch'])[:5], 1):
        if d['purch'] == 0 and d['spend'] == 0:
            continue
        roas = round(d['rev'] / d['spend'], 1) if d['spend'] else 0
        ctr  = round(d['clicks_all'] / d['imp'] * 100, 2) if d['imp'] else 0
        cpas_ad_rows.append(dict(rank=rank, name=name, short_name=short_name(name),
                                 spend=round(d['spend']), purch=round(d['purch']),
                                 rev=round(d['rev']), roas=roas, ctr=ctr))

    brand_total = engage_tot['spend'] + video_tot['spend'] + reach_tot['spend']
    return dict(
        branding=dict(engage=engage_tot, video=video_tot, reach=reach_tot,
                      targeting=tgt_rows, engage_ads=engage_ads, video_ads=video_ads,
                      total_spend=brand_total),
        messenger=dict(totals=msg_totals, campaigns=msg_camp_rows,
                       targeting=msg_tgt_rows, top_ads=msg_ad_rows),
        cpas=dict(totals=cpas_totals, adsets=cpas_adset_rows, top_ads=cpas_ad_rows),
        unclassified=sorted(unclassified),
    )


# ── Google Raw ────────────────────────────────────────────────────────────────

def parse_google_raw(ws):
    """
    Header row 3 (1-indexed), data from row 4.
    0-based: [80]=Campaign, [87]=Impressions, [88]=Interactions,
             [92]=Cost, [95/94]=Conversions
    GDN = 'gdn' in campaign name lowercase, else Search.
    """
    gdn    = defaultdict(lambda: dict(spend=0, imp=0, clicks=0, conv=0))
    search = defaultdict(lambda: dict(spend=0, imp=0, clicks=0, conv=0))

    _total_row = None

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) <= 92:
            continue
        camp = sv(row[80])
        # Capture dòng Total (camp='--' hoặc camp=None/empty sau khi strip)
        if not camp or camp == '--' or camp.lower().startswith('total'):
            if _total_row is None and (row[92] is not None and fv(row[92]) > 0):
                _total_row = row
            continue
        sp     = fv(row[92])
        imp    = fv_count(row[87])
        clicks = fv_count(row[88])
        conv   = fv(row[95]) if len(row) > 95 and row[95] is not None else (
                 fv(row[94]) if len(row) > 94 else 0)

        if 'gdn' in camp.lower():
            gdn[camp]['spend']  += sp
            gdn[camp]['imp']    += imp
            gdn[camp]['clicks'] += clicks
            gdn[camp]['conv']   += conv
        else:
            search[camp]['spend']  += sp
            search[camp]['imp']    += imp
            search[camp]['clicks'] += clicks
            search[camp]['conv']   += conv

    gdn_rows = []
    for name, d in sorted(gdn.items(), key=lambda x: -x[1]['spend']):
        cpm = round(d['spend'] / d['imp'] * 1000) if d['imp'] else 0
        gdn_rows.append(dict(name=name, spend=round(d['spend']),
                             imp=round(d['imp']), clicks=round(d['clicks']),
                             conv=round(d['conv'], 1), cpm=cpm))

    search_rows = []
    for name, d in sorted(search.items(), key=lambda x: -x[1]['spend']):
        ctr = round(d['clicks'] / d['imp'] * 100, 1) if d['imp'] else 0
        cpc = round(d['spend'] / d['clicks']) if d['clicks'] else 0
        search_rows.append(dict(name=name, spend=round(d['spend']), imp=round(d['imp']),
                                clicks=round(d['clicks']), ctr=ctr,
                                conv=round(d['conv'], 1), avg_cpc=cpc))

    if _total_row is not None:
        total_clicks = round(fv_count(_total_row[88])) if len(_total_row) > 88 else 0
        total_imp    = round(fv_count(_total_row[87])) if len(_total_row) > 87 else 0
    else:
        total_clicks = sum(d['clicks'] for d in gdn_rows) + sum(d['clicks'] for d in search_rows)
        total_imp    = sum(d['imp']    for d in gdn_rows) + sum(d['imp']    for d in search_rows)

    return dict(gdn=gdn_rows, search=search_rows,
                total_clicks=total_clicks, total_imp=total_imp)


# ── Google Keywords ───────────────────────────────────────────────────────────

def parse_google_kw(ws):
    """
    Header row 3, data from row 4.
    0-based: [1]=Keyword, [11]=Impr, [12]=Interactions, [15]=Cost,
             [23]=Conv, [24]=Avg CPC
    """
    keywords = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[1] is None:
            continue
        kw = sv(row[1])
        if not kw or kw.lower().startswith('total') or kw == '--':
            continue
        imp    = fv(row[11]); clicks = fv(row[12])
        cost   = fv(row[15]); conv   = fv(row[23])
        cpc    = fv(row[24]) if len(row) > 24 else (cost / clicks if clicks else 0)
        ctr    = round(clicks / imp * 100, 1) if imp else 0
        keywords.append(dict(kw=kw, imp=round(imp), clicks=round(clicks),
                             ctr=ctr, cost=round(cost), cpc=round(cpc), conv=round(conv, 1)))

    keywords.sort(key=lambda x: -x['clicks'])
    for i, kw in enumerate(keywords[:15], 1):
        kw['rank'] = i
    return keywords[:15]


# ── CPAS Plan sheet ───────────────────────────────────────────────────────────

def parse_cpas_plan(ws):
    """
    Sheet 'CPAS Plan': header row có tên tháng (Jan, Feb... May).
    Tìm cột 'May' động, đọc Revenue và Spending cho tháng đó.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # Find header row with month names + May column index
    may_col = None
    header_idx = None
    for i, row in enumerate(rows[:15]):
        for j, cell in enumerate(row):
            cs = sv(cell).lower().strip()
            if cs in ('may', 'tháng 5', 't5', '05', '5'):
                may_col = j
                header_idx = i
                break
        if may_col is not None:
            break

    if may_col is None:
        return {}

    # Scan rows below header for Revenue and Spending labels (col 0-3)
    revenue_val = 0
    spend_val   = 0
    for row in rows[header_idx + 1:]:
        if not row or may_col >= len(row):
            continue
        lbl = ''
        for col in range(min(4, len(row))):
            v = sv(row[col]).lower()
            if v:
                lbl = v
                break
        val = fv(row[may_col])
        if any(k in lbl for k in ('revenue', 'doanh thu', 'gmv', 'doanh số', 'rev')):
            revenue_val = val
        elif any(k in lbl for k in ('spend', 'budget', 'chi phí', 'ngân sách', 'spending', 'cost')):
            spend_val = val

    roas_plan = round(revenue_val / spend_val, 1) if spend_val else 0
    return {
        'revenue': round(revenue_val),
        'budget':  round(spend_val),
        'roas_plan': roas_plan,
    }


# ── FB date range detection ───────────────────────────────────────────────────

def get_fb_date_range(ws):
    """Scan FB Raw sheet for date values (datetime object or string YYYY-MM-DD)."""
    dates = []
    for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
        if not row:
            continue
        for val in row[:3]:
            if isinstance(val, datetime.datetime):
                dates.append(val.date())
            elif isinstance(val, datetime.date):
                dates.append(val)
            elif isinstance(val, str):
                try:
                    dates.append(datetime.date.fromisoformat(val.strip()))
                except (ValueError, AttributeError):
                    pass
        if len(dates) >= 20:
            break
    if dates:
        d_start, d_end = min(dates), max(dates)
        days = (d_end - d_start).days + 1
        return {'start': d_start, 'end': d_end, 'days': days}
    return {}


# ── CRM — Messenger Sale Data ─────────────────────────────────────────────────

def parse_crm(ws, date_start=None, date_end=None):
    """
    Header row 5 (1-indexed), data from row 7.
    0-based columns:
      [0]=STT, [1]=Ngày, [2]=Sản phẩm, [3]=Giá bán, [4]=SL,
      [7]=Tổng cộng, [8]=Tình trạng đơn hàng, [13]=Cũ/mới,
      [14]=Kênh
    date_start/date_end: filter theo tháng đầy đủ của report (first–last day of month).
    """
    orders = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or not isinstance(row[0], (int, float)) or row[1] is None:
            continue
        status  = sv(row[8])
        channel = sv(row[14])
        product = sv(row[2])
        kh_type = sv(row[13])
        date    = row[1]
        raw_rev = row[7]
        revenue = fv(raw_rev)
        # Cells with formula results store value in thousands unit (e.g. 535.5 = 535,500đ)
        if isinstance(raw_rev, (int, float)) and not isinstance(raw_rev, bool) and revenue < 100_000:
            revenue *= 1000
        qty     = fv(row[4])

        if date_start and date_end:
            order_date = date.date() if isinstance(date, datetime.datetime) else (
                date if isinstance(date, datetime.date) else None
            )
            if order_date is None or not (date_start <= order_date <= date_end):
                continue

        orders.append(dict(
            status=status, channel=channel, product=product,
            kh_type=kh_type, date=date, revenue=revenue, qty=qty
        ))

    delivered = [o for o in orders if 'giao' in o['status'].lower() and 'đang' not in o['status'].lower()]

    # Totals — tính tất cả đơn trong timeline (đã giao + đang giao)
    total_all_orders = len(orders)
    total_orders     = len(delivered)
    total_revenue    = sum(o['revenue'] for o in orders)
    total_qty        = sum(o['qty'] for o in orders)

    # Status funnel
    from collections import Counter
    status_counts = Counter(o['status'] for o in orders if o['status'])

    # Channel breakdown — tất cả đơn
    channel_rev = defaultdict(lambda: dict(orders=0, revenue=0))
    for o in orders:
        ch = o['channel'] or 'Unknown'
        channel_rev[ch]['orders']  += 1
        channel_rev[ch]['revenue'] += o['revenue']
    channel_rows = sorted(
        [dict(channel=k, **v) for k, v in channel_rev.items()],
        key=lambda x: -x['revenue']
    )

    # New vs returning — tất cả đơn
    kh_new = sum(1 for o in orders if 'mới' in o['kh_type'].lower())
    kh_old = sum(1 for o in orders if 'cũ' in o['kh_type'].lower())

    # Product breakdown — tất cả đơn
    prod_rev = defaultdict(lambda: dict(orders=0, revenue=0, qty=0))
    for o in orders:
        p = o['product'] or 'Unknown'
        prod_rev[p]['orders']  += 1
        prod_rev[p]['revenue'] += o['revenue']
        prod_rev[p]['qty']     += o['qty']
    product_rows = sorted(
        [dict(product=k, **v) for k, v in prod_rev.items()],
        key=lambda x: -x['revenue']
    )

    # Monthly trend — tất cả đơn
    monthly = defaultdict(lambda: dict(orders=0, revenue=0))
    for o in orders:
        if isinstance(o['date'], datetime.datetime):
            key = o['date'].strftime('%m/%Y')
            monthly[key]['orders']  += 1
            monthly[key]['revenue'] += o['revenue']
    monthly_rows = [dict(month=k, **v) for k, v in sorted(monthly.items())]

    # Bảng hiển thị: tất cả orders trong timeline (kể cả đang giao), sort theo ngày
    order_rows = []
    for o in sorted(orders, key=lambda x: x['date'] if isinstance(x['date'], datetime.datetime) else datetime.datetime.min):
        date_str = o['date'].strftime('%d/%m') if isinstance(o['date'], datetime.datetime) else sv(o['date'])
        order_rows.append(dict(
            date=date_str, product=o['product'], channel=o['channel'],
            kh_type=o['kh_type'], status=o['status'],
            revenue=o['revenue'], qty=o['qty'],
        ))

    return dict(
        total_all_orders=total_all_orders,
        total_orders=total_orders,
        total_revenue=round(total_revenue),
        total_qty=round(total_qty),
        kh_new=kh_new,
        kh_old=kh_old,
        status_funnel=dict(status_counts),
        channels=channel_rows,
        products=product_rows,
        monthly=monthly_rows,
        orders=order_rows,
    )


# ── Main entry point ──────────────────────────────────────────────────────────

def parse_all(file_bytes):
    """
    Parse single xlsx file with all sheets.
    Returns dict with: plan, branding, messenger, cpas, google, crm
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    plan_sheet      = find_sheet(wb, 'media plan') or find_sheet(wb, 'branding')
    fb_sheet        = find_sheet(wb, 'facebook raw')
    graw_sheet      = find_sheet(wb, 'google raw data') or find_sheet(wb, 'google raw')
    gkw_sheet       = find_sheet(wb, 'kw raw') or find_sheet(wb, 'google kw')
    crm_sheet       = find_sheet(wb, 'messenger sale') or find_sheet(wb, 'sale data')
    cpas_plan_sheet = find_sheet(wb, 'cpas plan') or find_sheet(wb, 'cpas')

    # Ensure graw is not the KW sheet
    if graw_sheet and 'kw' in graw_sheet.lower():
        graw_sheet = next(
            (s for s in wb.sheetnames if 'google raw' in s.lower() and 'kw' not in s.lower()),
            graw_sheet
        )

    result = {}

    if plan_sheet:
        result['plan'] = parse_media_plan(wb[plan_sheet])
    else:
        result['plan'] = {}

    date_range = {}
    if fb_sheet:
        date_range = get_fb_date_range(wb[fb_sheet])
        fb = parse_fb_raw(wb[fb_sheet])
        result['branding']   = fb['branding']
        result['messenger']  = fb['messenger']
        result['cpas']       = fb['cpas']
        result['unclassified'] = fb.get('unclassified', [])
    else:
        result['branding'] = result['messenger'] = result['cpas'] = {}
        result['unclassified'] = []
    result['date_range'] = date_range

    google_data = dict(gdn=[], search=[], keywords=[],
                       gdn_spend=0, search_spend=0, google_spend=0)
    if graw_sheet:
        graw = parse_google_raw(wb[graw_sheet])
        gdn_spend    = sum(d['spend'] for d in graw['gdn'])
        search_spend = sum(d['spend'] for d in graw['search'])
        google_data.update(dict(
            gdn=graw['gdn'], search=graw['search'],
            gdn_spend=gdn_spend, search_spend=search_spend,
            google_spend=gdn_spend + search_spend,
            total_clicks=graw['total_clicks'],
            total_imp=graw['total_imp'],
        ))
    if gkw_sheet:
        google_data['keywords'] = parse_google_kw(wb[gkw_sheet])

    result['google'] = google_data

    if crm_sheet:
        result['crm'] = parse_crm(
            wb[crm_sheet],
            date_start=date_range.get('start'),
            date_end=date_range.get('end'),
        )
    else:
        result['crm'] = {}

    result['cpas_plan'] = parse_cpas_plan(wb[cpas_plan_sheet]) if cpas_plan_sheet else {}
    result['_sheets']   = wb.sheetnames

    # Overall totals
    brand_sp = result.get('branding', {}).get('total_spend', 0)
    msg_sp   = result.get('messenger', {}).get('totals', {}).get('spend', 0)
    cpas_sp  = result.get('cpas', {}).get('totals', {}).get('spend', 0)
    gg_sp    = result['google']['google_spend']
    result['overall'] = dict(
        total_spend=round(brand_sp + msg_sp + cpas_sp + gg_sp),
        branding_spend=round(brand_sp),
        messenger_spend=round(msg_sp),
        cpas_spend=round(cpas_sp),
        gdn_spend=round(result['google']['gdn_spend']),
        search_spend=round(result['google']['search_spend']),
    )

    return result
